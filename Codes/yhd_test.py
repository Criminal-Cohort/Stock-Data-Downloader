import tkinter as tk
import os
import time
import openpyxl
import yfinance as yh
import datetime
import pandas as pd
import csv


root = tk.Tk()
root.title("Yahoo Downloader")

# Set window size
root.geometry("500x450")

# Date input section with "Custom Date" label
date_frame = tk.Frame(root)
date_frame.pack(pady=20)

custom_date_label = tk.Label(date_frame, text="Custom Date", font=("Arial", 14))
custom_date_label.pack(side="left", padx=10)

# Year input
year_entry = tk.Entry(date_frame, width=6, font=("Arial", 14))
year_entry.pack(side="left")

# Slash label
slash_label1 = tk.Label(date_frame, text="/", font=("Arial", 14))
slash_label1.pack(side="left")

# Month input
month_entry = tk.Entry(date_frame, width=4, font=("Arial", 14))
month_entry.pack(side="left")

# Slash label
slash_label2 = tk.Label(date_frame, text="/", font=("Arial", 14))
slash_label2.pack(side="left")

# Day input
day_entry = tk.Entry(date_frame, width=4, font=("Arial", 14))
day_entry.pack(side="left")

# Ticker input section
ticker_frame = tk.Frame(root)
ticker_frame.pack(pady=10)

ticker_label = tk.Label(ticker_frame, text="Ticker:", font=("Arial", 14))
ticker_label.pack(side="left")

ticker_entry = tk.Entry(ticker_frame, font=("Arial", 14), width=15)
ticker_entry.pack(side="left")


try:
    path_file = os.path.join(os.getcwd(), "Book1.xlsx")
    path = os.path.join(os.getcwd(), "historicalData")

    wb_obj = openpyxl.load_workbook(path_file)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row

except Exception as e:
    print(e)
    print("Book1 Not Found!")
    time.sleep(10)
    quit()


def download_data(file, start, end):
    download_source = os.path.join(path, f"{file}.csv")
    df = yh.download(file, start=start, end=end)
    df.to_csv(download_source)
    print(df)
    progress_text.insert(tk.END, df)
    progress_text.insert(tk.END, "\n\n*********************************************************************************************\n")
    converter(download_source)


def extract_files():
    file_names = []
    for i in range(1, max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        file_name = str(cell_obj.value).upper()
        file_names.append(file_name)
    return file_names


def process_files(file_names, start, end):
    for file_name in file_names:
        try:
            download_data(file_name, start, end)

        except Exception as e:
            print(f"{file_name} Not Found!")


def yes_custom():
    try:
        file = ticker_entry.get()
        dty = int(year_entry.get())
        dtm = int(month_entry.get())
        dtd = int(day_entry.get())

        start = datetime.datetime(dty, dtm, dtd)
        end = datetime.datetime.today()

        download_data(file, start, end)

    except Exception as e:
        print(e)
        progress_text.insert(tk.END, str(e))


def yes_normal():
    dt = datetime.datetime.now().replace(year=datetime.datetime.now().year - 10)
    dtd, dtm = dt.day, dt.month

    if dtd == 29 and dtm == 2:
        dtd, dtm = 1, 3

    try:
        file = ticker_entry.get()
        start = datetime.datetime(dt.year, dtm, dtd)
        end = datetime.datetime.today()

        download_data(file, start, end)

    except Exception as e:
        print(e)
        progress_text.insert(tk.END, str(e))


def no():
    dt = datetime.datetime.now().replace(year=datetime.datetime.now().year - 10)
    dtd, dtm = dt.day, dt.month

    if dtd == 29 and dtm == 2:
        dtd, dtm = 1, 3

    file_names = extract_files()
    start = datetime.datetime(dt.year, dtm, dtd)
    end = datetime.datetime.today()

    process_files(file_names, start, end)


def yes():
    dty = int(year_entry.get())
    dtm = int(month_entry.get())
    dtd = int(day_entry.get())

    file_names = extract_files()
    start = datetime.datetime(dty - 10, dtm, dtd)
    end = datetime.datetime(dty, dtm, dtd)

    process_files(file_names, start, end)


def main_cust():
    try:
        if ticker_entry.get():
            if year_entry.get() and month_entry.get() and day_entry.get():
                yes_custom()

            else:
                yes_normal()

        else:
            if year_entry.get() and month_entry.get() and day_entry.get():
                yes()

            else:
                no()

    except Exception as e:
        no()
        print(e)


def converter(file):
    col_list = ["Date", "Open", "High", "Low", "Close", "Adj Close", "Volume"]
    df = pd.read_csv(file, usecols=col_list)

    data_reversed = {
        "timestamp": df["Date"][::-1],
        "open": df["Open"][::-1],
        "high": df["High"][::-1],
        "low": df["Low"][::-1],
        "close": df["Close"][::-1],
        "adjusted_close": df["Adj Close"][::-1],
        "volume": df["Volume"][::-1],
        "dividend": [0] * len(df),
        "split_coefficient": [1] * len(df)
    }

    with open(file, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(data_reversed.keys())
        writer.writerows(zip(*data_reversed.values()))


# Start button
start_button = tk.Button(root, text="Start", font=("Arial", 14), width=10, command=main_cust)
start_button.pack(pady=20)

progress_text = tk.Text(root, height=410, width=320, font=("Arial", 12))
progress_text.pack(pady=10)

root.after(10000, main_cust)

root.mainloop()