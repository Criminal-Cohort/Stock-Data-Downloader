import os
import datetime
import openpyxl
import yfinance as yh
import time
from inputimeout import inputimeout, TimeoutOccurred
import pandas as pd
import csv


try:
    path_file = os.path.join(os.getcwd(), "Book1.xlsx")
    path = os.path.join(os.getcwd(), "historicalData")

    wb_obj = openpyxl.load_workbook(path_file)
    sheet_obj = wb_obj.active
    max_row = sheet_obj.max_row

except Exception as e:
    print(e)
    print("Book1 Not Found!")
    time.sleep(10)
    quit()


def download_data(file, start, end):
    download_source = os.path.join(path, f"{file}.csv")
    df = yh.download(file, start=start, end=end)
    df.to_csv(download_source)
    print(df)
    converter(download_source)


def extract_files():
    file_names = []
    for i in range(1, max_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        file_name = str(cell_obj.value).upper()
        file_names.append(file_name)
    return file_names


def process_files(file_names, start, end):
    for file_name in file_names:
        try:
            download_data(file_name, start, end)

        except Exception as e:
            print(f"{file_name} Not Found!")


def yes_custom():
    try:
        file = input("Please Enter Name: ")
        dty = int(input("Please enter your custom year: "))
        dtm = int(input("Please enter your custom month: "))
        dtd = int(input("Please enter your custom day: "))

        start = datetime.datetime(dty, dtm, dtd)
        end = datetime.datetime.today()

        download_data(file, start, end)

    except Exception as e:
        print(e)


def yes_normal():
    dt = datetime.datetime.now().replace(year=datetime.datetime.now().year - 10)
    dtd, dtm = dt.day, dt.month

    if dtd == 29 and dtm == 2:
        dtd, dtm = 1, 3

    try:
        file = input("Please Enter Name: ")
        start = datetime.datetime(dt.year, dtm, dtd)
        end = datetime.datetime.today()

        download_data(file, start, end)

    except Exception as e:
        print(e)


def no():
    dt = datetime.datetime.now().replace(year=datetime.datetime.now().year - 10)
    dtd, dtm = dt.day, dt.month

    if dtd == 29 and dtm == 2:
        dtd, dtm = 1, 3

    file_names = extract_files()
    start = datetime.datetime(dt.year, dtm, dtd)
    end = datetime.datetime.today()

    process_files(file_names, start, end)


def yes():
    dty = int(input("Please enter your custom year: "))
    dtm = int(input("Please enter your custom month: "))
    dtd = int(input("Please enter your custom day: "))

    file_names = extract_files()
    start = datetime.datetime(dty - 10, dtm, dtd)
    end = datetime.datetime(dty, dtm, dtd)

    process_files(file_names, start, end)


def main():
    try:
        single = inputimeout(prompt="Do you want a single file? (Y/N): ", timeout=10)
        if single.lower() in ["y", "yes"]:
            yes_normal()

        elif single.lower() in ["n", "no"]:
            custom = inputimeout(prompt="Do you want a custom date? (Y/N): ", timeout=10)
            if custom.lower() in ["y", "yes"]:
                yes()

            else:
                no()

    except TimeoutOccurred:
        no()


def converter(file):
    col_list = ["Date", "Open", "High", "Low", "Close", "Adj Close", "Volume"]
    df = pd.read_csv(file, usecols=col_list)

    data_reversed = {
        "timestamp": df["Date"][::-1],
        "open": df["Open"][::-1],
        "high": df["High"][::-1],
        "low": df["Low"][::-1],
        "close": df["Close"][::-1],
        "adjusted_close": df["Adj Close"][::-1],
        "volume": df["Volume"][::-1],
        "dividend": [0] * len(df),
        "split_coefficient": [1] * len(df)
    }

    with open(file, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(data_reversed.keys())
        writer.writerows(zip(*data_reversed.values()))


main()